# -*- coding: utf-8 -*-
import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

files = [
    ('ARDC PP Final report.xlsx', 'ARDC PP'),
    ('ARDC Sales Final report.xlsx', 'ARDC Sales'),
    ('ENF and GF final report.xlsx', 'ENF/GF')
]

all_data = {}

for filepath, name in files:
    wb = openpyxl.load_workbook(filepath)

    # Count items
    obtained = list(wb['Obtained Items'].iter_rows(min_row=6, values_only=True))
    obtained_count = len([r for r in obtained if r[0] is not None])

    missing = list(wb['Missing Items'].iter_rows(min_row=6, values_only=True))
    missing_count = len([r for r in missing if r[0] is not None])

    findings = list(wb['Additional Findings'].iter_rows(min_row=6, values_only=True))
    findings_count = len([r for r in findings if r[0] is not None])

    high_risk = len([r for r in findings if r[0] is not None and r[2] == 'high'])

    print(f'{name}: Obtained={obtained_count}, Missing={missing_count}, Findings={findings_count}, High Risk={high_risk}')

    # Collect detailed findings
    all_data[name] = {
        'obtained': obtained_count,
        'missing': missing_count,
        'findings': findings_count,
        'high_risk': high_risk,
        'findings_detail': []
    }

    for r in findings:
        if r[0] is not None and r[2] == 'high':
            all_data[name]['findings_detail'].append({
                'topic': str(r[0])[:100] if r[0] else '',
                'type': str(r[1]) if r[1] else '',
                'risk': str(r[2]) if r[2] else '',
                'details': str(r[3])[:300] if r[3] else '',
                'sap_analysis': str(r[4])[:300] if r[4] else '',
                'recommendation': str(r[5])[:300] if r[5] else '',
            })

# Save to JSON
with open('analysis_results.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2, ensure_ascii=False)

print("\nData saved to analysis_results.json")
