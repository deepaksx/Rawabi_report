# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Get key obtained items from each file
print("=" * 80)
print("KEY OBTAINED INFORMATION")
print("=" * 80)

# ARDC PP - Obtained Items
print("\n### ARDC - PRODUCTION (PP) ###")
wb = openpyxl.load_workbook('ARDC PP Final report.xlsx')
ws = wb['Obtained Items']
for row in ws.iter_rows(min_row=6, values_only=True):
    if row[0] is not None and row[2] == 'critical':
        item = str(row[1])[:60] if row[1] else ''
        info = str(row[4])[:200] if row[4] else ''
        print(f"- {item}")
        print(f"  Info: {info}...")
        print()

# ARDC Sales - Obtained Items
print("\n### ARDC - SALES (SD) ###")
wb = openpyxl.load_workbook('ARDC Sales Final report.xlsx')
ws = wb['Obtained Items']
for row in ws.iter_rows(min_row=6, values_only=True):
    if row[0] is not None and row[2] == 'critical':
        item = str(row[1])[:60] if row[1] else ''
        info = str(row[4])[:200] if row[4] else ''
        print(f"- {item}")
        print(f"  Info: {info}...")
        print()

# ENF/GF - Obtained Items
print("\n### ENF & GF - POULTRY OPERATIONS ###")
wb = openpyxl.load_workbook('ENF and GF final report.xlsx')
ws = wb['Obtained Items']
for row in ws.iter_rows(min_row=6, values_only=True):
    if row[0] is not None and row[2] == 'critical':
        item = str(row[1])[:60] if row[1] else ''
        info = str(row[4])[:200] if row[4] else ''
        print(f"- {item}")
        print(f"  Info: {info}...")
        print()

# Get missing items
print("\n" + "=" * 80)
print("KEY MISSING INFORMATION (Not Obtained)")
print("=" * 80)

print("\n### ARDC - PP MISSING ###")
wb = openpyxl.load_workbook('ARDC PP Final report.xlsx')
ws = wb['Missing Items']
count = 0
for row in ws.iter_rows(min_row=6, values_only=True):
    if row[0] is not None and row[2] == 'critical' and count < 10:
        item = str(row[1])[:80] if row[1] else ''
        print(f"- {item}")
        count += 1

print("\n### ARDC - SD MISSING ###")
wb = openpyxl.load_workbook('ARDC Sales Final report.xlsx')
ws = wb['Missing Items']
count = 0
for row in ws.iter_rows(min_row=6, values_only=True):
    if row[0] is not None and row[2] == 'critical' and count < 10:
        item = str(row[1])[:80] if row[1] else ''
        print(f"- {item}")
        count += 1
